# ========= フリーズ防止のおまじない =========
import matplotlib
matplotlib.use('Agg')
# ==========================================

import streamlit as st
import pulp
import pandas as pd
import matplotlib.pyplot as plt
import japanize_matplotlib  # ★追加：クラウドでも日本語を表示させる魔法
from datetime import datetime, timedelta
import random

# Mac標準の日本語フォントを指定（★この1行を消すか、先頭に # をつけてコメントアウトします）
# plt.rcParams['font.family'] = 'Hiragino Sans'

st.set_page_config(page_title="PIGC シフト自動作成アプリ", layout="wide")

# ==========================================
# ⚙️ 左側のサイドバー（設定画面）の構築
# ==========================================
st.sidebar.header("⚙️ 基本設定")

# 1. 日付の設定
st.sidebar.subheader("📅 期間の設定")
start_date = st.sidebar.date_input("シフト開始日", value=datetime(2026, 3, 21))
num_days = st.sidebar.number_input("期間（日数）", min_value=7, max_value=60, value=31)


# 2. 全スタッフの登録（★改行で入力できるように大改良！）
st.sidebar.subheader("👤 全スタッフの登録")
st.sidebar.caption("※1行に1人ずつ名前を入力してください（Enterで改行）\n※エクセル等からコピー＆ペーストも可能です")

default_staffs = "小畑 明美\n柴田 美幸\n河本 陽平\n植木 健太郎\n深田 尚\n佐藤 大吾郎\n中谷 吉男\n浜崎 圭子"
master_staff_str = st.sidebar.text_area("全スタッフ名", default_staffs, height=230)

# 入力された文字列を「改行」で分割してリスト化
all_staffs = [s.strip() for s in master_staff_str.split("\n") if s.strip()]


# 3. ポジションの振り分け（上で登録したメンバーから選択）
st.sidebar.subheader("🏢 ポジション設定")
st.sidebar.caption("※上で登録したメンバーから選んでください")

staffs_am = st.sidebar.multiselect(
    "午前スタッフ", 
    options=all_staffs, 
    default=[s for s in ["小畑 明美", "柴田 美幸", "河本 陽平", "植木 健太郎"] if s in all_staffs]
)

staffs_pm = st.sidebar.multiselect(
    "午後スタッフ", 
    options=all_staffs, 
    default=[s for s in ["深田 尚", "佐藤 大吾郎", "中谷 吉男", "浜崎 圭子"] if s in all_staffs]
)

cross_staffs = st.sidebar.multiselect(
    "午前・午後 両方入るスタッフ (通しあり)", 
    options=all_staffs, 
    default=[s for s in ["小畑 明美", "柴田 美幸"] if s in all_staffs]
)


# 4. NGペアの設定（データエディタで選択式）
st.sidebar.subheader("🆖 NGペア設定")
st.sidebar.caption("※同じ日に入れないペアを選択・追加できます")

df_ng_init = pd.DataFrame([{"スタッフ1": "河本 陽平", "スタッフ2": "植木 健太郎"}])

edited_ng = st.sidebar.data_editor(
    df_ng_init,
    num_rows="dynamic",
    hide_index=True,
    column_config={
        "スタッフ1": st.column_config.SelectboxColumn("スタッフ1", options=all_staffs),
        "スタッフ2": st.column_config.SelectboxColumn("スタッフ2", options=all_staffs),
    }
)

ng_pairs = []
for _, row in edited_ng.iterrows():
    s1, s2 = row.get("スタッフ1"), row.get("スタッフ2")
    if pd.notna(s1) and pd.notna(s2) and s1 != s2 and s1 in all_staffs and s2 in all_staffs:
        ng_pairs.append((s1, s2))


# 5. 出勤日数の条件設定（データエディタ）
st.sidebar.subheader("📊 出勤日数の条件")
st.sidebar.caption("※表の数字をクリックして書き換えられます（0は条件なし）")

default_min_dict = {"小畑 明美": 20, "柴田 美幸": 15, "植木 健太郎": 13, "深田 尚": 10, "中谷 吉男": 10, "佐藤 大吾郎": 8}
default_exact_dict = {"河本 陽平": 5, "浜崎 圭子": 0}

df_rules = pd.DataFrame({
    "スタッフ": all_staffs,
    "最低出勤日数": [default_min_dict.get(s, 0) for s in all_staffs],
    "ぴったり日数": [default_exact_dict.get(s, 0) for s in all_staffs]
})

edited_rules = st.sidebar.data_editor(df_rules, hide_index=True)

min_shifts = {row["スタッフ"]: int(row["最低出勤日数"]) for _, row in edited_rules.iterrows() if int(row["最低出勤日数"]) > 0}
exact_shifts = {row["スタッフ"]: int(row["ぴったり日数"]) for _, row in edited_rules.iterrows() if int(row["ぴったり日数"]) > 0}


# ==========================================
# 共通データ（日付リスト）の生成
# ==========================================
days_map = ["月", "火", "水", "木", "金", "土", "日"]
dates_info = []

for i in range(num_days): 
    current_date = start_date + timedelta(days=i)
    date_str = f"{current_date.month}/{current_date.day}"
    day_str = days_map[current_date.weekday()]
    
    is_sat = current_date.weekday() == 5
    is_sun = current_date.weekday() == 6
    is_weekend = is_sat or is_sun
    
    dates_info.append({
        "id": date_str, 
        "display_am": f"{day_str}\n{date_str}", 
        "display_pm": f"{date_str}\n{day_str}", 
        "is_weekend": is_weekend,
        "is_sat": is_sat,
        "is_sun": is_sun,
        "day_str": day_str
    })

date_ids = [d["id"] for d in dates_info]


# 設定が変わった場合に備えて、状態（gui_states）をリセット・更新する処理
if 'gui_states' not in st.session_state or st.session_state.get('last_staffs') != all_staffs or st.session_state.get('last_dates') != date_ids:
    st.session_state.gui_states = {s: {d: 0 for d in date_ids} for s in all_staffs}
    st.session_state.last_staffs = all_staffs
    st.session_state.last_dates = date_ids


def get_val(var):
    val = pulp.value(var)
    return val if val is not None else 0

# ==========================================
# ファイル出力処理（人数変動に動的対応）
# ==========================================
def generate_files(x_am, x_pm, shain_am_vars, shain_pm_vars, status_box):
    schedule_data = []
    
    header_row = ["曜日\n日付"] + staffs_am + ["日付\n曜日"] + staffs_pm + ["午前", "午後"]
    
    title_row = [""] * len(header_row)
    title_row[1] = "【 午前打席 】"
    if len(staffs_am) + 2 < len(title_row):
        title_row[len(staffs_am) + 2] = "【 午後打席 】"
    title_row[-2] = "【 社員 】"
    
    schedule_data.append(title_row)
    schedule_data.append(header_row)
    
    for d_info in dates_info:
        d_id = d_info["id"]
        row = [d_info["display_am"]]
        
        for s in staffs_am:
            am_val = get_val(x_am[s, d_id])
            pm_val = get_val(x_pm[s, d_id])
            if am_val == 1 and pm_val == 1:
                row.append("●☆") 
            elif am_val == 1:
                row.append("●")  
            else:
                row.append("")
            
        row.append(d_info["display_pm"])
        
        for s in staffs_pm:
            pm_val = get_val(x_pm[s, d_id])
            row.append("●" if pm_val == 1 else "") 
            
        am_help = get_val(shain_am_vars[d_id])
        pm_help = get_val(shain_pm_vars[d_id])
        row.append("●" if am_help > 0.5 else "") 
        row.append("●" if pm_help > 0.5 else "") 
        
        schedule_data.append(row)
        
    row_total = ["出勤日数"]
    for s in staffs_am:
        row_total.append(str(int(sum(get_val(x_am[s, d]) for d in date_ids))))
    row_total.append("出勤日数")
    for s in staffs_pm:
        row_total.append(str(int(sum(get_val(x_pm[s, d]) for d in date_ids))))
        
    shain_am_sum = int(sum(get_val(shain_am_vars[d]) for d in date_ids))
    shain_pm_sum = int(sum(get_val(shain_pm_vars[d]) for d in date_ids))
    row_total.extend([str(shain_am_sum) if shain_am_sum > 0 else "0", 
                      str(shain_pm_sum) if shain_pm_sum > 0 else "0"])
    schedule_data.append(row_total)
    
    row_wh = ["平日/休日"]
    for s in staffs_am:
        w_days = sum(1 for d in dates_info if get_val(x_am[s, d['id']]) == 1 and not d['is_weekend'])
        h_days = sum(1 for d in dates_info if get_val(x_am[s, d['id']]) == 1 and d['is_weekend'])
        row_wh.append(f"{w_days}\n{h_days}")
    row_wh.append("平日/休日")
    for s in staffs_pm:
        w_days = sum(1 for d in dates_info if get_val(x_pm[s, d['id']]) == 1 and not d['is_weekend'])
        h_days = sum(1 for d in dates_info if get_val(x_pm[s, d['id']]) == 1 and d['is_weekend'])
        row_wh.append(f"{w_days}\n{h_days}")
        
    shain_am_w = sum(1 for d in dates_info if get_val(shain_am_vars[d['id']]) > 0 and not d['is_weekend'])
    shain_am_h = sum(1 for d in dates_info if get_val(shain_am_vars[d['id']]) > 0 and d['is_weekend'])
    shain_pm_w = sum(1 for d in dates_info if get_val(shain_pm_vars[d['id']]) > 0 and not d['is_weekend'])
    shain_pm_h = sum(1 for d in dates_info if get_val(shain_pm_vars[d['id']]) > 0 and d['is_weekend'])
    row_wh.extend([f"{shain_am_w}\n{shain_am_h}", f"{shain_pm_w}\n{shain_pm_h}"])
    schedule_data.append(row_wh)

    row_pm_count = ["通し"]
    for s in staffs_am:
        if s in cross_staffs:
            pm_sum = int(sum(get_val(x_pm[s, d]) for d in date_ids))
            row_pm_count.append(str(pm_sum) if pm_sum > 0 else "")
        else:
            row_pm_count.append("")
    row_pm_count.append("")
    for s in staffs_pm:
        row_pm_count.append("")
    row_pm_count.extend(["", ""])
    schedule_data.append(row_pm_count)
    
    df = pd.DataFrame(schedule_data)
    
    status_box.info("⏳ 【ステップ2/3】エクセルファイルを作成中...")
    excel_filename = "full_shift_schedule.xlsx"
    df.to_excel(excel_filename, index=False, header=False)
    try:
        import openpyxl
        from openpyxl.styles import Border, Side, Alignment
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        thin_border = Side(style='thin', color='000000')
        
        if len(staffs_am) > 0:
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+len(staffs_am))
            ws.cell(row=1, column=2).alignment = Alignment(horizontal='center')
        
        pm_col_start = len(staffs_am) + 3
        if len(staffs_pm) > 0:
            pm_col_end = pm_col_start + len(staffs_pm) - 1
            ws.merge_cells(start_row=1, start_column=pm_col_start, end_row=1, end_column=pm_col_end)
            ws.cell(row=1, column=pm_col_start).alignment = Alignment(horizontal='center')
        else:
            pm_col_end = pm_col_start - 1
        
        shain_col_start = pm_col_end + 1
        shain_col_end = shain_col_start + 1
        ws.merge_cells(start_row=1, start_column=shain_col_start, end_row=1, end_column=shain_col_end)
        ws.cell(row=1, column=shain_col_start).alignment = Alignment(horizontal='center')
        
        for row in ws.iter_rows(min_row=1, max_row=len(schedule_data)):
            for cell in row:
                cell.border = Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)
                
        wb.save(excel_filename)
        wb.close()
    except Exception as e:
        pass
    
    status_box.info("⏳ 【ステップ3/3】写真(画像)ファイルを作成中...")
    fig_width = max(14, len(all_staffs) * 1.5)
    fig, ax = plt.subplots(figsize=(fig_width, 13))
    ax.axis('off')
    
    table = ax.table(cellText=df.values, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.0, 1.6) 
    
    for (row_idx, col_idx), cell in table.get_celld().items():
        if row_idx == 0:
            cell.set_linewidth(0) 
            cell.set_fontsize(12)
            cell.set_text_props(weight='bold')
        elif row_idx == 1:
            cell.set_facecolor('#d9edf7') 
            cell.set_text_props(weight='bold')
        elif 2 <= row_idx <= 2 + num_days - 1:
            d_info = dates_info[row_idx - 2]
            if d_info["is_weekend"]:
                cell.set_facecolor('#e6f2ff')
            if col_idx in [0, len(staffs_am) + 1]: 
                if d_info["is_sat"]:
                    cell.get_text().set_color('blue')
                elif d_info["is_sun"]:
                    cell.get_text().set_color('red')
        elif row_idx >= len(schedule_data) - 4:
            cell.set_facecolor('#fff2cc') 
            cell.set_text_props(weight='bold')
            
    fig.canvas.draw()
    plt.tight_layout()
    png_filename = "full_shift_schedule.png"
    plt.savefig(png_filename, dpi=300, bbox_inches='tight')
    plt.close('all')
    
    return excel_filename, png_filename


# ==========================================
# 最適化の実行
# ==========================================
def run_optimization(status_box):
    status_box.info("⏳ 【ステップ1/3】シフトの最適な組み合わせをAIで計算中...（最大10秒）")
    
    prob = pulp.LpProblem("Full_Shift_Scheduling", pulp.LpMinimize)
    
    x_am = {}
    x_pm = {}
    for s in all_staffs:
        for d in date_ids:
            x_am[s, d] = pulp.LpVariable(f"xam_{s}_{d}", cat="Binary")
            x_pm[s, d] = pulp.LpVariable(f"xpm_{s}_{d}", cat="Binary")

    shain_am_vars = {d: pulp.LpVariable(f"shain_am_{d}", cat="Integer", lowBound=0) for d in date_ids}
    shain_pm_vars = {d: pulp.LpVariable(f"shain_pm_{d}", cat="Integer", lowBound=0) for d in date_ids}

    prob += (
        pulp.lpSum(1000 * shain_am_vars[d] + 1000 * shain_pm_vars[d] for d in date_ids) +
        pulp.lpSum(random.random() * x_am[s, d] for s in all_staffs for d in date_ids) +
        pulp.lpSum(random.random() * x_pm[s, d] for s in all_staffs for d in date_ids)
    )

    for d in date_ids:
        prob += pulp.lpSum(x_am[s, d] for s in all_staffs) + shain_am_vars[d] == 2
        prob += pulp.lpSum(x_pm[s, d] for s in all_staffs) + shain_pm_vars[d] == 2

    for s in all_staffs:
        for d in date_ids:
            state = st.session_state.gui_states[s][d]
            if state == 1: 
                prob += x_am[s, d] == 0
                prob += x_pm[s, d] == 0
            elif state == 2: 
                prob += x_pm[s, d] == 0
                
            if s in cross_staffs:
                prob += x_pm[s, d] <= x_am[s, d]
            elif s in staffs_am and s not in staffs_pm:
                prob += x_pm[s, d] == 0
            elif s in staffs_pm and s not in staffs_am:
                prob += x_am[s, d] == 0
            elif s not in staffs_am and s not in staffs_pm:
                prob += x_am[s, d] == 0
                prob += x_pm[s, d] == 0

    for p1, p2 in ng_pairs:
        if p1 in all_staffs and p2 in all_staffs:
            for d in date_ids:
                prob += x_am[p1, d] + x_am[p2, d] <= 1
                prob += x_pm[p1, d] + x_pm[p2, d] <= 1

    for s in all_staffs:
        work_days = pulp.lpSum(x_am[s, d] for d in date_ids) if s in staffs_am else pulp.lpSum(x_pm[s, d] for d in date_ids)
        if s in min_shifts:
            prob += work_days >= min_shifts[s]
        if s in exact_shifts:
            prob += work_days == exact_shifts[s]

    for s in all_staffs:
        for i in range(len(date_ids) - 6):
            if s in staffs_am:
                prob += pulp.lpSum(x_am[s, date_ids[i+j]] for j in range(7)) <= 6
            else:
                prob += pulp.lpSum(x_pm[s, date_ids[i+j]] for j in range(7)) <= 6

    solver = pulp.PULP_CBC_CMD(timeLimit=10, msg=False)
    status = prob.solve(solver)

    if pulp.LpStatus[status] == "Optimal" or status == 1:
        return generate_files(x_am, x_pm, shain_am_vars, shain_pm_vars, status_box)
    else:
        return None, None


# ==========================================
# 📺 Web画面メイン部分
# ==========================================
st.title("⛳️ PIGC シフト自動作成アプリ")
st.write("左側のサイドバーで基本設定を行い、下のタブで各スタッフの出勤できない日を設定してください。")

def toggle_state(staff, d_id):
    current = st.session_state.gui_states[staff][d_id]
    if staff in cross_staffs:
        if current == 0: st.session_state.gui_states[staff][d_id] = 2
        elif current == 2: st.session_state.gui_states[staff][d_id] = 1
        else: st.session_state.gui_states[staff][d_id] = 0
    else:
        if current == 0: st.session_state.gui_states[staff][d_id] = 1
        else: st.session_state.gui_states[staff][d_id] = 0

tabs = st.tabs(all_staffs)

for i, staff in enumerate(all_staffs):
    with tabs[i]:
        if staff in cross_staffs:
            st.info(f"💡 {staff} さん（「🟠 午前」は午後は入れない日の設定です）")
        else:
            st.info(f"💡 {staff} さんの出勤できない日を設定してください")
            
        for row_idx in range((num_days // 7) + 1): 
            cols = st.columns(7)
            for col_idx in range(7):
                day_idx = row_idx * 7 + col_idx
                if day_idx < len(dates_info):
                    d_info = dates_info[day_idx]
                    d_id = d_info["id"]
                    state = st.session_state.gui_states[staff][d_id]
                    
                    if staff in cross_staffs:
                        if state == 0: text = f"🟢 両方\n{d_id}"
                        elif state == 2: text = f"🟠 午前\n{d_id}"
                        else: text = f"❌ 休み\n{d_id}"
                    else:
                        if state == 0: text = f"🟢 出勤\n{d_id}"
                        else: text = f"❌ 休み\n{d_id}"
                    
                    cols[col_idx].button(text, key=f"btn_{staff}_{d_id}", on_click=toggle_state, args=(staff, d_id), use_container_width=True)

st.divider()

if st.button("✨ この条件でシフト表を作成する", type="primary", use_container_width=True):
    status_box = st.empty()
    excel_file, png_file = run_optimization(status_box)
        
    if excel_file and png_file:
        status_box.success("🎉 シフトの作成に成功しました！以下のボタンからダウンロードできます。")
        
        col1, col2 = st.columns(2)
        with col1:
            with open(excel_file, "rb") as f:
                st.download_button(label="📗 エクセル版をダウンロード", data=f, file_name="shift_schedule.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col2:
            with open(png_file, "rb") as f:
                st.download_button(label="🖼️ 写真(PNG)版をダウンロード", data=f, file_name="shift_schedule.png", mime="image/png", use_container_width=True)
                
        st.image(png_file, caption="完成したシフト表プレビュー")
    else:
        status_box.error("⚠️ 指定された休みが多すぎるか、条件が厳しすぎてシフトが組めませんでした。条件をゆるめて再度お試しください。")
